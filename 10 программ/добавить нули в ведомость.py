import openpyxl
from openpyxl.styles import Font
from tkinter import filedialog
import tkinter as tk
import os

def run_macro():
    file_path = file_path_entry.get()
    start_row = int(start_row_entry.get())
    end_row = int(end_row_entry.get())
    from_column = from_column_entry.get()
    to_column = to_column_entry.get()

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Создаем стиль для белого цвета шрифта Arial размером 11
    white_font = Font(color='00FFFFFF', name='Arial', size=11)

    for row in range(start_row, end_row + 1):
        cell_value = ws[f'V{row}'].value
        if isinstance(cell_value, (int, float)):
            for col in range(ord(from_column) - 65, ord(to_column) - 64):  # Преобразуем букву столбца в номер
                if ws.cell(row=row, column=col + 1).value is None:
                    ws.cell(row=row, column=col + 1, value=0)
                    ws.cell(row=row, column=col + 1).font = white_font  # Применяем стиль для белого цвета шрифта Arial размером 11

    wb.save(file_path)
    result_label.config(text="Макрос успешно выполнен. Результат сохранен в выбранном файле.")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    file_path_entry.delete(0, tk.END)
    file_path_entry.insert(0, file_path)

# Создаем окно
root = tk.Tk()
root.title("Добавление белых нулей")
root.option_add('*Font', 'Arial 22')

# Создаем элементы управления
file_label = tk.Label(root, text="Выберите файл:")
file_label.pack()

file_path_entry = tk.Entry(root, width=50, font=('Arial', 22))
file_path_entry.pack()

browse_button = tk.Button(root, text="Обзор", command=browse_file, font=('Arial', 22))
browse_button.pack()

start_row_label = tk.Label(root, text="Первая строка для проверки:", font=('Arial', 22))
start_row_label.pack()

start_row_entry = tk.Entry(root, width=10, font=('Arial', 22))
start_row_entry.pack()

end_row_label = tk.Label(root, text="Последняя строка для проверки:", font=('Arial', 22))
end_row_label.pack()

end_row_entry = tk.Entry(root, width=10, font=('Arial', 22))
end_row_entry.pack()

from_column_label = tk.Label(root, text="Столбец от (A, B, C, ...):", font=('Arial', 22))
from_column_label.pack()

from_column_entry = tk.Entry(root, width=5, font=('Arial', 22))
from_column_entry.pack()

to_column_label = tk.Label(root, text="Столбец до (A, B, C, ...):", font=('Arial', 22))
to_column_label.pack()

to_column_entry = tk.Entry(root, width=5, font=('Arial', 22))
to_column_entry.pack()

run_button = tk.Button(root, text="Выполнить", command=run_macro, font=('Arial', 22))
run_button.pack()

result_label = tk.Label(root, text="", font=('Arial', 22))
result_label.pack()

root.mainloop()
