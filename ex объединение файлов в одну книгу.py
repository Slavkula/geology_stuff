import openpyxl
import os

# Путь к папке с файлами Excel
folder_path = 'C:/Users/1/Desktop/описание 2 часть/'

# Создаем новый файл Excel для объединения данных
combined_wb = openpyxl.Workbook()

# Читаем каждый файл и добавляем его данные в общий файл Excel
for file in os.listdir(folder_path):
    if file.endswith('.xlsx'):
        file_path = os.path.join(folder_path, file)  # Полный путь к файлу
        wb = openpyxl.load_workbook(file_path)  # Загружаем книгу Excel

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]  # Выбираем лист
            combined_sheet = combined_wb.create_sheet(title=file.split('.xlsx')[0])  # Создаем лист в общем файле с именем файла
            for row in sheet.iter_rows(values_only=True):
                combined_sheet.append(row)  # Копируем данные из листа в общий файл

combined_wb.save('объединенный_файл.xlsx')

# листы в файле будут называться как называются исходные файлы
