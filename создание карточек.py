import openpyxl
import tkinter as tk
from tkinter import filedialog
from tkinter import font
import os

def process_excel_data(file_path, template_file, start_row, end_row, cells, output_cells, filename_cells):
    wb2 = openpyxl.load_workbook(file_path)
    sheet2 = wb2.active

    for i in range(start_row, end_row + 1):
        data = [sheet2[f'{cell}{i}'].value if cell else None for cell in cells]

        filename_part1 = sheet2[f'{filename_cells[0]}{i}'].value if filename_cells[0] else None
        filename_part2 = sheet2[f'{filename_cells[1]}{i}'].value if filename_cells[1] else None

        wb_new = openpyxl.load_workbook(template_file)
        sheet_new = wb_new.active

        for j in range(len(output_cells)):
            if data[j] is not None:
                if j == 0:
                    sheet_new[output_cells[j]] = "Скважина №" + str(data[j])
                else:
                    sheet_new[output_cells[j]] = data[j]

        if filename_part1 and filename_part2:
            wb_new.save(f'{filename_part1}-{filename_part2}.xlsx')

def on_process_button_click():
    start_row = int(entry_start.get())
    end_row = int(entry_end.get())
    cells = [entry.get() for entry in entries]
    output_cells = [entry.get() for entry in output_entries]
    filename_cells = [entry.get() for entry in filename_entries]
    process_excel_data(data_file_path, card_file_path, start_row, end_row, cells, output_cells, filename_cells)
    label_result.config(text="Данные успешно обработаны и сохранены в новых файлах.")

def on_data_file_button_click():
    global data_file_path
    data_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if data_file_path:
        data_file_button.config(bg="green")

def on_card_file_button_click():
    global card_file_path
    card_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if card_file_path:
        card_file_button.config(bg="green")

window = tk.Tk()
window.title("Создание карточек")

# Increase the font size
default_font = font.nametofont("TkDefaultFont")
default_font.configure(size=20)

frame1 = tk.Frame(window)
frame1.pack(pady=10)

label_data_file = tk.Label(frame1, text="Файл ведомости:")
label_data_file.pack(side=tk.LEFT)
data_file_button = tk.Button(frame1, text="Выбрать файл", command=on_data_file_button_click)
data_file_button.pack(side=tk.LEFT, padx=10)

label_card_file = tk.Label(frame1, text="Файл карточки:")
label_card_file.pack(side=tk.LEFT)
card_file_button = tk.Button(frame1, text="Выбрать файл", command=on_card_file_button_click)
card_file_button.pack(side=tk.LEFT, padx=10)

frame2 = tk.Frame(window)
frame2.pack(pady=10)

label_start = tk.Label(frame2, text="Начальная строка:")
label_start.pack(side=tk.LEFT)
entry_start = tk.Entry(frame2)
entry_start.pack(side=tk.LEFT, padx=10)

label_end = tk.Label(frame2, text="Конечная строка:")
label_end.pack(side=tk.LEFT)
entry_end = tk.Entry(frame2)
entry_end.pack(side=tk.LEFT, padx=10)

frame3 = tk.Frame(window)
frame3.pack(pady=10)

entries = []
for i in range(10):
    if i == 0:
        label = tk.Label(frame3, text=f"Столбец {i+1}:")
    else:
        label = tk.Label(frame3, text=f"{i+1}:")
    label.pack(side=tk.LEFT)
    entry = tk.Entry(frame3, width=5)
    entry.pack(side=tk.LEFT, padx=2)
    entries.append(entry)

frame4 = tk.Frame(window)
frame4.pack(pady=10)

output_entries = []
for i in range(10):
    if i == 0:
        label = tk.Label(frame4, text=f"Ячейка {i+1}:")
    else:
        label = tk.Label(frame4, text=f"{i+1}:")
    label.pack(side=tk.LEFT)
    entry = tk.Entry(frame4, width=5)
    entry.pack(side=tk.LEFT, padx=2)
    output_entries.append(entry)

frame5 = tk.Frame(window)
frame5.pack(pady=10)

filename_entries = []
for i in range(2):
    label = tk.Label(frame5, text=f"Ячейка имени файла {i+1}:")
    label.pack(side=tk.LEFT)
    entry = tk.Entry(frame5, width=3)
    entry.pack(side=tk.LEFT, padx=10)
    filename_entries.append(entry)

process_button = tk.Button(window, text="Создать карточки", command=on_process_button_click)
process_button.pack()

label_result = tk.Label(window, text="")
label_result.pack()

# Add instructions to the bottom of the window
instructions = tk.Label(window, text="""
Инструкция:
1. Строки: это с какой брать данные и до какой
Если тебе нужна только одна, то указываешь число два раза.
2. Столбцы это из каких брать (Например А).
3. Ячейки это в какие копировать (например А1)
Первая ячейка всегда записывается со словом 'Скважина №'
Если она не нужна, то начинаешь со вторых.
4. Столбцы для создания имени файла (Только буква).
Имя файла будет 'Первая-вторая', берутся из ведомости.
5. Объединить нужные файлы в одну книгу можно кнопками ниже.
""")
instructions.pack()


def select_files():
    global file_paths
    file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    files_label.config(text="Выбрано " + str(len(file_paths)) + " файлов")

def merge_files():
    if file_paths:
        combined_wb = openpyxl.Workbook()
        for file_path in file_paths:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                combined_sheet = combined_wb.create_sheet(title=os.path.basename(file_path).split('.xlsx')[0])
                for row in sheet.iter_rows(values_only=True):
                    combined_sheet.append(row)

        first_sheet = combined_wb.sheetnames[0]
        if first_sheet == "Sheet":
            del combined_wb[first_sheet]

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            combined_wb.save(save_path)

# Создаем контейнер для кнопок
button_frame = tk.Frame(window)
button_frame.pack(pady=10)

# Создаем виджеты в контейнере
select_files_button = tk.Button(button_frame, text="Выбрать файлы", command=select_files)
select_files_button.pack(side="left", padx=5)

merge_button = tk.Button(button_frame, text="Объединить файлы", command=merge_files)
merge_button.pack(side="left", padx=5)

# Создаем виджет для отображения количества выбранных файлов
files_label = tk.Label(window, text="Выбрано 0 файлов")
files_label.pack()

# Запускаем главный цикл обработки событий
window.mainloop()
